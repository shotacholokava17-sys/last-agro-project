import os
import io
import sqlite3
import socket
import shutil
from collections import defaultdict
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from PIL import Image
import imagehash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "plant_secret_key_123"

UPLOAD_FOLDER = os.path.join('static', 'uploads')
REF_FOLDER = os.path.join('static', 'references')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REF_FOLDER, exist_ok=True)


# ==================== მონაცემთა ბაზის ინიციალიზაცია ====================
def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()
    # დაავადებები და მათი ფასების პარამეტრები
    conn.execute('''
        CREATE TABLE IF NOT EXISTS disease_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            multiplier REAL DEFAULT 1.0,
            cost REAL DEFAULT 1.0,
            description TEXT DEFAULT ''
        )
    ''')
    # migration: description სვეტის დამატება თუ ძველ ბაზაში არ არის
    try:
        conn.execute("ALTER TABLE disease_settings ADD COLUMN description TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass  # სვეტი უკვე არსებობს
    # საცნობარო სურათები საძიებლად
    conn.execute('''
        CREATE TABLE IF NOT EXISTS reference_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            disease_name TEXT,
            image_path TEXT,
            hash_val TEXT
        )
    ''')
    # მიმდინარე ან ძველი კვლევის სესიები
    conn.execute('''
        CREATE TABLE IF NOT EXISTS research_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            status TEXT DEFAULT 'active' -- 'active' ან 'archived'
        )
    ''')
    # ატვირთული მცენარეების ანალიზის შედეგები
    conn.execute('''
        CREATE TABLE IF NOT EXISTS plant_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER,
            position TEXT,
            image_path TEXT,
            detected_disease TEXT,
            calculated_cost REAL,
            created_at TEXT,
            FOREIGN KEY (session_id) REFERENCES research_sessions(id)
        )
    ''')

    # საწყისი დაავადებების ჩაყრა, თუ ბაზა ცარიელია
    default_diseases = [
        ("ჯანსაღი", 0.0, 0.0),
        ("ჭრაქი", 8.0, 0.5),
        ("სიდამპლე", 3.0, 0.7),
        ("ნაცარი", 3.0, 0.7)
    ]
    for name, mult, cost in default_diseases:
        conn.execute('INSERT OR IGNORE INTO disease_settings (name, multiplier, cost) VALUES (?, ?, ?)',
                     (name, mult, cost))

    # თუ აქტიური სესია არ არსებობს, შევქმნათ
    active_session = conn.execute("SELECT id FROM research_sessions WHERE status = 'active'").fetchone()
    if not active_session:
        conn.execute("INSERT INTO research_sessions (created_at, status) VALUES (?, 'active')",
                     (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))

    conn.commit()
    conn.close()


init_db()


# ==================== გამოსახულების ჰეშირება და შედარება ====================
def calculate_hash(img_path):
    try:
        with Image.open(img_path) as img:
            return str(imagehash.phash(img))
    except Exception as e:
        print(f"ჰეშის შეცდომა: {e}")
        return None


def classify_image(img_path):
    img_hash_str = calculate_hash(img_path)
    if not img_hash_str:
        return "შეცდომა"

    img_hash = imagehash.hex_to_hash(img_hash_str)

    conn = get_db_connection()
    refs = conn.execute("SELECT disease_name, hash_val FROM reference_images").fetchall()
    conn.close()

    if not refs:
        return "უცნობი (საცნობარო ბაზა ცარიელია)"

    best_match = "უცნობი"
    min_dist = 100

    for ref in refs:
        ref_hash = imagehash.hex_to_hash(ref['hash_val'])
        dist = img_hash - ref_hash
        if dist < min_dist:
            min_dist = dist
            best_match = ref['disease_name']

    return best_match if min_dist < 12 else "უცნობი"


def classify_by_filename(filename, diseases):
    """ფაილის სახელით ამოიცნობს დაავადებას"""
    filename_lower = filename.lower()
    for d in diseases:
        if d['name'].lower() in filename_lower:
            return d['name']
    return None


def auto_save_to_references(conn, disease_name, img_path, hash_val):
    """ავტომატურად ინახავს ფოტოს საცნობარო ბაზაში, თუ ჰეში უნიკალურია"""
    existing = conn.execute("SELECT id FROM reference_images WHERE hash_val = ?", (hash_val,)).fetchone()
    if existing:
        return  # უკვე არსებობს

    # ფოტოს კოპირება references საქაღალდეში
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    original_name = os.path.basename(img_path)
    ref_filename = f"ref_auto_{timestamp}_{original_name}"
    ref_filepath = os.path.join(REF_FOLDER, ref_filename)
    shutil.copy2(os.path.join('static', img_path), ref_filepath)

    conn.execute("INSERT INTO reference_images (disease_name, image_path, hash_val) VALUES (?, ?, ?)",
                 (disease_name, f"references/{ref_filename}", hash_val))


def group_records_by_position(records):
    """ჩანაწერები დაჯგუფებს პოზიციების მიხედვით"""
    groups = defaultdict(list)
    for r in records:
        groups[r['position']].append(r)
    # გარდაქმნა sorted list-ად რიცხვითი მნიშვნელობის მიხედვით
    def extract_number(pos):
        if pos and pos.startswith("მცენარე "):
            try:
                return int(pos.split(" ")[1])
            except (ValueError, IndexError):
                pass
        return 0
    return sorted(groups.items(), key=lambda x: extract_number(x[0]))


# ==================== IP-ს გაგება ტელეფონისთვის ====================
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip


# ==================== როუტები ====================

@app.route('/')
def index():
    conn = get_db_connection()
    active_session = conn.execute(
        "SELECT * FROM research_sessions WHERE status = 'active' ORDER BY id DESC LIMIT 1").fetchone()

    records = []
    total_cost = 0.0
    grouped_records = []

    if active_session:
        records = conn.execute("""
            SELECT r.*, s.multiplier, s.cost 
            FROM plant_records r 
            LEFT JOIN disease_settings s ON r.detected_disease = s.name 
            WHERE r.session_id = ? ORDER BY r.position ASC, r.id ASC
        """, (active_session['id'],)).fetchall()

        grouped_records = group_records_by_position(records)

        sum_row = conn.execute("SELECT SUM(calculated_cost) as total FROM plant_records WHERE session_id = ?",
                               (active_session['id'],)).fetchone()
        if sum_row['total']:
            total_cost = sum_row['total']

    diseases = conn.execute("SELECT name FROM disease_settings").fetchall()
    disease_descriptions = {
        row['name']: row['description']
        for row in conn.execute("SELECT name, description FROM disease_settings").fetchall()
    }
    conn.close()

    return render_template('index.html',
                           session=active_session,
                           records=records,
                           grouped_records=grouped_records,
                           total_cost=total_cost,
                           diseases=diseases,
                           disease_descriptions=disease_descriptions)


@app.route('/upload', methods=['POST'])
def upload():
    if 'photos' not in request.files:
        flash("ფაილები ვერ მოიძებნა", "danger")
        return redirect(url_for('index'))

    files = request.files.getlist('photos')
    files = [f for f in files if f.filename != '']
    if not files:
        flash("ფაილები ცარიელია", "danger")
        return redirect(url_for('index'))

    conn = get_db_connection()

    # ===== ყოველი ატვირთვა = ახალი სესია =====
    # 1. არსებული active სესია გადადის archived-ში
    conn.execute("UPDATE research_sessions SET status = 'archived' WHERE status = 'active'")
    # 2. ახალი სესია იქმნება
    conn.execute("INSERT INTO research_sessions (created_at, status) VALUES (?, 'active')",
                 (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
    conn.commit()

    # ახლახანს შექმნილი სესიის ID
    active_session_id = conn.execute("SELECT id FROM research_sessions WHERE status = 'active'").fetchone()['id']

    uploaded_count = 0
    diseases = conn.execute("SELECT name FROM disease_settings").fetchall()

    for idx, file in enumerate(files, 1):
        position = f"მცენარე {idx}"

        # ფაილის შენახვა მიკროწამებით (collision-ის თავიდან ასაცილებლად)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"plant_{timestamp}_{file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        # ===== 1. ფაილის სახელით ამოცნობა =====
        detected_disease = classify_by_filename(file.filename, diseases)

        if detected_disease:
            h_val = calculate_hash(filepath)
            if h_val:
                auto_save_to_references(conn, detected_disease, f"uploads/{filename}", h_val)
        else:
            # ===== 2. ჰეშით ამოცნობა =====
            detected_disease = classify_image(filepath)

        # ===== ახლად ამოცნობილი (არა "უცნობი") — ავტომატურად ვინახავთ საცნობარო ბაზაში =====
        if detected_disease not in ["უცნობი", "შეცდომა", "უცნობი (საცნობარო ბაზა ცარიელია)"]:
            h_val = calculate_hash(filepath)
            if h_val:
                auto_save_to_references(conn, detected_disease, f"uploads/{filename}", h_val)

        setting = conn.execute("SELECT multiplier, cost FROM disease_settings WHERE name = ?",
                               (detected_disease,)).fetchone()
        calculated_cost = (setting['multiplier'] * setting['cost']) if setting else 0.0

        conn.execute("""
            INSERT INTO plant_records (session_id, position, image_path, detected_disease, calculated_cost, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (active_session_id, position, f"uploads/{filename}", detected_disease, calculated_cost,
              datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

        uploaded_count += 1

    conn.commit()
    conn.close()

    flash(f"✅ კვლევა #{active_session_id} — {uploaded_count} ფოტო გაანალიზდა!", "success")
    return redirect(url_for('index'))


@app.route('/assign_disease/<int:record_id>', methods=['POST'])
def assign_disease(record_id):
    new_disease = request.form.get('disease_name')
    if not new_disease:
        flash("დაავადება არ მითითებულა!", "danger")
        return redirect(url_for('index'))

    conn = get_db_connection()
    record = conn.execute("SELECT * FROM plant_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        flash("ჩანაწერი ვერ მოიძებნა!", "danger")
        conn.close()
        return redirect(url_for('index'))

    # ახალი ზარალის გამოთვლა
    setting = conn.execute("SELECT multiplier, cost FROM disease_settings WHERE name = ?",
                           (new_disease,)).fetchone()
    calculated_cost = (setting['multiplier'] * setting['cost']) if setting else 0.0

    # ბაზის განახლება
    conn.execute("UPDATE plant_records SET detected_disease = ?, calculated_cost = ? WHERE id = ?",
                 (new_disease, calculated_cost, record_id))

    # ======= ვინახავთ საცნობარო ბაზაში სწავლებისთვის =======
    h_val = calculate_hash(os.path.join('static', record['image_path']))
    if h_val:
        auto_save_to_references(conn, new_disease, record['image_path'], h_val)

    conn.commit()

    # ვამოწმებთ საიდან მოვიდა მოთხოვნა (history თუ index)
    redirect_to = request.form.get('redirect_to', 'index')
    conn.close()

    flash(f"✅ დაავადება განახლდა: {new_disease} — სისტემამ ისწავლა ეს ფოტო!", "success")
    if redirect_to == 'history':
        return redirect(url_for('history'))
    return redirect(url_for('index'))


@app.route('/new_session', methods=['POST'])
def new_session():
    conn = get_db_connection()
    conn.execute("UPDATE research_sessions SET status = 'archived' WHERE status = 'active'")
    conn.execute("INSERT INTO research_sessions (created_at, status) VALUES (?, 'active')",
                 (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
    conn.commit()
    conn.close()
    flash("ახალი კვლევის ეტაპი წარმატებით დაიწყო! წინა მონაცემები გადავიდა ისტორიაში.", "success")
    return redirect(url_for('index'))


# ==================== Excel ექსპორტი ====================
@app.route('/export_excel/<int:session_id>')
def export_excel(session_id):
    conn = get_db_connection()
    session_info = conn.execute("SELECT * FROM research_sessions WHERE id = ?", (session_id,)).fetchone()
    records = conn.execute("""
        SELECT r.*, s.multiplier, s.cost 
        FROM plant_records r
        LEFT JOIN disease_settings s ON r.detected_disease = s.name
        WHERE r.session_id = ? ORDER BY r.position ASC, r.id ASC
    """, (session_id,)).fetchall()
    total = conn.execute("SELECT SUM(calculated_cost) as total FROM plant_records WHERE session_id = ?",
                         (session_id,)).fetchone()['total'] or 0
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"კვლევა #{session_id}"

    # ===================== სტილები =====================
    header_fill = PatternFill("solid", fgColor="1e3a2f")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill("solid", fgColor="2e7d32")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    healthy_fill = PatternFill("solid", fgColor="d4edda")
    unknown_fill = PatternFill("solid", fgColor="e2e3e5")
    disease_fill = PatternFill("solid", fgColor="fff3cd")
    total_fill = PatternFill("solid", fgColor="ffeeba")
    total_font = Font(bold=True, size=11, color="856404")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===================== სათაური =====================
    ws.merge_cells('A1:G1')
    ws['A1'] = f"🌿 AgroDetector — კვლევა #{session_id}"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:G2')
    session_date = session_info['created_at'] if session_info else "-"
    ws['A2'] = f"📅 ჩატარების თარიღი: {session_date}"
    ws['A2'].font = Font(italic=True, color="FFFFFF", size=10)
    ws['A2'].fill = header_fill
    ws['A2'].alignment = center_align

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # ===================== სვეტების სათაური =====================
    headers = ["#", "📍 პოზიცია", "🦠 დიაგნოზი", "💰 ზარალი (₾)", "📊 კოეფ.", "💵 ბაზ. ფასი", "🕐 დრო"]
    col_widths = [5, 28, 20, 16, 10, 12, 18]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[3].height = 22

    # ===================== მონაცემები =====================
    current_row = 4
    grouped = group_records_by_position(records)

    for pos_idx, (position, pos_records) in enumerate(grouped):
        # პოზიციის სათაური
        ws.merge_cells(f'A{current_row}:G{current_row}')
        pos_cell = ws.cell(row=current_row, column=1, value=f"📍 {position}")
        pos_cell.font = Font(bold=True, color="1e3a2f", size=10)
        pos_cell.fill = PatternFill("solid", fgColor="c8e6c9")
        pos_cell.alignment = left_align
        pos_cell.border = thin_border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for idx, r in enumerate(pos_records, 1):
            disease = r['detected_disease']
            if disease == 'ჯანსაღი':
                row_fill = healthy_fill
            elif disease == 'უცნობი':
                row_fill = unknown_fill
            else:
                row_fill = disease_fill

            data = [
                idx,
                r['position'],
                disease,
                round(r['calculated_cost'], 2),
                r['multiplier'] if r['multiplier'] is not None else "-",
                r['cost'] if r['cost'] is not None else "-",
                r['created_at']
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = center_align if col_num != 2 else left_align
            ws.row_dimensions[current_row].height = 16
            current_row += 1

    # ===================== ჯამი =====================
    ws.merge_cells(f'A{current_row}:C{current_row}')
    total_label = ws.cell(row=current_row, column=1, value="💰 ჯამური ზარალი:")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border = thin_border

    total_val = ws.cell(row=current_row, column=4, value=round(total, 2))
    total_val.font = total_font
    total_val.fill = total_fill
    total_val.alignment = center_align
    total_val.border = thin_border

    for col in range(5, 8):
        c = ws.cell(row=current_row, column=col)
        c.fill = total_fill
        c.border = thin_border

    ws.row_dimensions[current_row].height = 22

    # ===================== გაგზავნა =====================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    export_filename = f"AgroDetector_kvleva_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=export_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== დაავადებების ბაზის მართვა ====================
@app.route('/references', methods=['GET', 'POST'])
def references():
    conn = get_db_connection()
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_ref':
            disease_name = request.form.get('disease_name')
            file = request.files.get('ref_photo')
            if file and disease_name:
                filename = f"ref_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                filepath = os.path.join(REF_FOLDER, filename)
                file.save(filepath)

                h_val = calculate_hash(filepath)
                if h_val:
                    conn.execute("INSERT INTO reference_images (disease_name, image_path, hash_val) VALUES (?, ?, ?)",
                                 (disease_name, f"references/{filename}", h_val))
                    conn.commit()
                    flash(f"ახალი ფოტო დაემატა დაავადებას: {disease_name}", "success")
                else:
                    flash("სურათის ჰეშირება ვერ მოხერხდა!", "danger")

        elif action == 'update_rates':
            for key, value in request.form.items():
                if key.startswith('mult_'):
                    dis_name = key.replace('mult_', '')
                    mult = float(value)
                    cost = float(request.form.get(f'cost_{dis_name}', 1.0))
                    desc = request.form.get(f'desc_{dis_name}', '').strip()
                    conn.execute(
                        "UPDATE disease_settings SET multiplier = ?, cost = ?, description = ? WHERE name = ?",
                        (mult, cost, desc, dis_name)
                    )
            conn.commit()
            flash("ფასების კოეფიციენტები წარმატებით განახლდა!", "success")

        elif action == 'add_disease_type':
            new_name = request.form.get('new_disease_name')
            if new_name:
                conn.execute("INSERT OR IGNORE INTO disease_settings (name, multiplier, cost) VALUES (?, 1.0, 1.0)",
                             (new_name,))
                conn.commit()
                flash(f"დაავადების ტიპი '{new_name}' დაემატა!", "success")

    settings = conn.execute("SELECT * FROM disease_settings").fetchall()
    refs = conn.execute("SELECT * FROM reference_images").fetchall()
    conn.close()
    return render_template('references.html', settings=settings, refs=refs)


# ==================== წაშლის ახალი ფუნქციონალი ====================

@app.route('/delete_reference/<int:ref_id>', methods=['POST'])
def delete_reference(ref_id):
    conn = get_db_connection()
    ref = conn.execute("SELECT * FROM reference_images WHERE id = ?", (ref_id,)).fetchone()
    if ref:
        # ფიზიკური ფაილის წაშლა დისკიდან
        full_path = os.path.join('static', ref['image_path'])
        if os.path.exists(full_path):
            os.remove(full_path)

        conn.execute("DELETE FROM reference_images WHERE id = ?", (ref_id,))
        conn.commit()
        flash("საცნობარო ფოტო წარმატებით წაიშალა!", "success")
    conn.close()
    return redirect(url_for('references'))


@app.route('/delete_disease/<string:disease_name>', methods=['POST'])
def delete_disease(disease_name):
    conn = get_db_connection()

    # 1. ჯერ ვშლით ამ დაავადების ყველა საცნობარო ფოტოს დისკიდან
    refs = conn.execute("SELECT * FROM reference_images WHERE disease_name = ?", (disease_name,)).fetchall()
    for ref in refs:
        full_path = os.path.join('static', ref['image_path'])
        if os.path.exists(full_path):
            os.remove(full_path)

    # 2. ვშლით საცნობარო ცხრილიდან
    conn.execute("DELETE FROM reference_images WHERE disease_name = ?", (disease_name,))
    # 3. ვშლით ძირითადი პარამეტრების ცხრილიდან
    conn.execute("DELETE FROM disease_settings WHERE name = ?", (disease_name,))

    conn.commit()
    conn.close()
    flash(f"დაავადება '{disease_name}' და მისი ყველა საცნობარო ფოტო წაიშალა!", "success")
    return redirect(url_for('references'))


@app.route('/delete_session/<int:session_id>', methods=['POST'])
def delete_session(session_id):
    conn = get_db_connection()

    # 1. ამ სესიის ყველა სურათის წაშლა დისკიდან
    records = conn.execute("SELECT * FROM plant_records WHERE session_id = ?", (session_id,)).fetchall()
    for r in records:
        full_path = os.path.join('static', r['image_path'])
        if os.path.exists(full_path):
            os.remove(full_path)

    # 2. წავშალოთ ჩანაწერები ბაზიდან
    conn.execute("DELETE FROM plant_records WHERE session_id = ?", (session_id,))
    conn.execute("DELETE FROM research_sessions WHERE id = ?", (session_id,))

    conn.commit()
    conn.close()
    flash(f"კვლევის ეტაპი #{session_id} და მისი ფოტოები წარმატებით წაიშალა!", "success")
    return redirect(url_for('history'))


# ===================================================================

@app.route('/history')
def history():
    conn = get_db_connection()
    # ყველა სესია — მიმდინარეც და archived-იც, უახლესი ზემოთ
    sessions = conn.execute("SELECT * FROM research_sessions ORDER BY id DESC").fetchall()
    diseases = conn.execute("SELECT name FROM disease_settings").fetchall()

    archived_data = []
    for s in sessions:
        records = conn.execute("""
            SELECT r.*, ds.multiplier, ds.cost 
            FROM plant_records r
            LEFT JOIN disease_settings ds ON r.detected_disease = ds.name
            WHERE r.session_id = ? ORDER BY r.position ASC, r.id ASC
        """, (s['id'],)).fetchall()
        total = conn.execute("SELECT SUM(calculated_cost) as total FROM plant_records WHERE session_id = ?",
                             (s['id'],)).fetchone()['total'] or 0
        grouped = group_records_by_position(records)
        archived_data.append({
            'session': s,
            'records': records,
            'grouped_records': grouped,
            'total_cost': total,
            'is_active': s['status'] == 'active'
        })

    conn.close()
    return render_template('history.html', archived_data=archived_data, diseases=diseases)


@app.route('/plants')
def plants():
    conn = get_db_connection()
    rows = conn.execute("SELECT DISTINCT position FROM plant_records ORDER BY position ASC").fetchall()
    
    plant_summaries = []
    for row in rows:
        pos = row['position']
        stats = conn.execute("""
            SELECT COUNT(*) as scan_count, SUM(calculated_cost) as total_loss
            FROM plant_records WHERE position = ?
        """, (pos,)).fetchone()
        
        latest = conn.execute("""
            SELECT detected_disease, created_at, image_path 
            FROM plant_records WHERE position = ? ORDER BY id DESC LIMIT 1
        """, (pos,)).fetchone()
        
        plant_summaries.append({
            'position': pos,
            'scan_count': stats['scan_count'],
            'total_loss': stats['total_loss'] or 0.0,
            'latest_disease': latest['detected_disease'] if latest else 'უცნობი',
            'latest_date': latest['created_at'] if latest else '-',
            'latest_image': latest['image_path'] if latest else None
        })
    
    conn.close()
    return render_template('plants.html', plants=plant_summaries)


@app.route('/plants/<path:position>')
def plant_detail(position):
    conn = get_db_connection()
    records = conn.execute("""
        SELECT r.*, s.multiplier, s.cost, rs.status as session_status
        FROM plant_records r
        LEFT JOIN disease_settings s ON r.detected_disease = s.name
        LEFT JOIN research_sessions rs ON r.session_id = rs.id
        WHERE r.position = ? ORDER BY r.id DESC
    """, (position,)).fetchall()
    
    if not records:
        conn.close()
        flash("მცენარე ამ პოზიციით ვერ მოიძებნა!", "warning")
        return redirect(url_for('plants'))
        
    stats = conn.execute("""
        SELECT COUNT(*) as scan_count, SUM(calculated_cost) as total_loss
        FROM plant_records WHERE position = ?
    """, (position,)).fetchone()
    
    diseases = conn.execute("SELECT name FROM disease_settings").fetchall()
    conn.close()
    
    return render_template('plant_detail.html', 
                           position=position, 
                           records=records, 
                           scan_count=stats['scan_count'], 
                           total_loss=stats['total_loss'] or 0.0,
                           diseases=diseases)


if __name__ == '__main__':
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    local_ip = get_local_ip()
    print("\n" + "=" * 60)
    print(f"AgroDetector - Web Interface started!")
    print(f"Mobile: http://{local_ip}:5000")
    print(f"PC:     http://127.0.0.1:5000")
    print("=" * 60 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=True)